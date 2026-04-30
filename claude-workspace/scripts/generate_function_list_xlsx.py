"""
Generate clinic_management_function_list.xlsx from function_list_data.py.

Usage:
    python scripts/generate_function_list_xlsx.py

Output:
    docs/clinic_management_function_list.xlsx
"""

import sys
from collections import Counter
from pathlib import Path

# Add scripts dir to path
SCRIPT_DIR = Path(__file__).parent
sys.path.insert(0, str(SCRIPT_DIR))

from function_list_data import FUNCTIONS, DONE, WIP, TODO, IDEA  # noqa: E402

import openpyxl  # noqa: E402
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402
from openpyxl.worksheet.table import Table, TableStyleInfo  # noqa: E402

OUTPUT_PATH = SCRIPT_DIR.parent.parent / "docs" / "clinic_management_function_list.xlsx"

# ──────────────────────────────────────────────────────────────────────
# Styling
# ──────────────────────────────────────────────────────────────────────
BRAND_BLUE = "1A6BAC"
BRAND_LIGHT = "EEF4FB"
HEADER_TEXT = "FFFFFF"

STATUS_COLORS = {
    DONE: "C6EFCE",  # light green
    WIP: "FFEB9C",   # light yellow
    TODO: "F2F2F2",  # light gray
    IDEA: "FCE4D6",  # light orange
}

GROUP_COLORS = {
    "AUTH": "E8F4FD", "RBAC": "E8F4FD",
    "TENANT": "FCE4D6", "SUB": "FCE4D6",
    "PATIENT": "DEEBF7", "VISIT": "DEEBF7", "VITAL": "DEEBF7", "DIAG": "DEEBF7",
    "SVC": "E2EFDA", "RX": "E2EFDA", "MED": "E2EFDA", "PHRM": "E2EFDA",
    "APPT": "FFF2CC", "BILL": "FFF2CC", "HR": "FFF2CC",
    "RPT": "FCE4D6", "NOTI": "FCE4D6",
    "AUDIT": "F4CCCC", "CFG": "F4CCCC",
    "PLT": "EAD1DC",
    "JOB": "D9D2E9", "DATA": "D9D2E9", "INT": "D9D2E9",
    "I18N": "D0E0E3", "A11Y": "D0E0E3", "THEME": "D0E0E3",
}

PHASE_COLORS = {
    "v1": "C6EFCE",
    "v2": "FFEB9C",
    "v3": "FCE4D6",
}

THIN_BORDER = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)


# ──────────────────────────────────────────────────────────────────────
# Build Functions sheet
# ──────────────────────────────────────────────────────────────────────
def build_functions_sheet(wb):
    ws = wb.active
    ws.title = "Functions"

    headers = [
        "Nhóm", "Mã", "Tên chức năng", "Mô tả ngắn", "Mô tả chi tiết",
        "Role", "Phase", "Task", "Status",
    ]
    ws.append(headers)

    # Format header row
    header_font = Font(name="Calibri", size=11, bold=True, color=HEADER_TEXT)
    header_fill = PatternFill("solid", fgColor=BRAND_BLUE)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = THIN_BORDER
    ws.row_dimensions[1].height = 32

    # Data rows
    for row_idx, fn in enumerate(FUNCTIONS, start=2):
        ws.cell(row=row_idx, column=1, value=fn["group"])
        ws.cell(row=row_idx, column=2, value=fn["code"])
        ws.cell(row=row_idx, column=3, value=fn["name"])
        ws.cell(row=row_idx, column=4, value=fn["brief"])
        ws.cell(row=row_idx, column=5, value=fn["detail"])
        ws.cell(row=row_idx, column=6, value=fn["role"])
        ws.cell(row=row_idx, column=7, value=fn["phase"])
        ws.cell(row=row_idx, column=8, value=fn["task"])
        ws.cell(row=row_idx, column=9, value=fn["status"])

        # Apply formatting
        group_color = GROUP_COLORS.get(fn["group"], "FFFFFF")
        ws.cell(row=row_idx, column=1).fill = PatternFill("solid", fgColor=group_color)
        ws.cell(row=row_idx, column=1).font = Font(bold=True)

        ws.cell(row=row_idx, column=2).font = Font(name="Consolas", size=10)

        ws.cell(row=row_idx, column=3).font = Font(bold=True)

        ws.cell(row=row_idx, column=5).alignment = Alignment(wrap_text=True, vertical="top")

        phase_color = PHASE_COLORS.get(fn["phase"], "FFFFFF")
        ws.cell(row=row_idx, column=7).fill = PatternFill("solid", fgColor=phase_color)
        ws.cell(row=row_idx, column=7).alignment = Alignment(horizontal="center")

        status_color = STATUS_COLORS.get(fn["status"], "FFFFFF")
        ws.cell(row=row_idx, column=9).fill = PatternFill("solid", fgColor=status_color)
        ws.cell(row=row_idx, column=9).alignment = Alignment(horizontal="center")

        # All cells: top alignment + border + wrap
        for col_idx in range(1, 10):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = THIN_BORDER
            if cell.alignment.wrap_text is None:
                cell.alignment = Alignment(
                    horizontal=cell.alignment.horizontal or "left",
                    vertical="top",
                    wrap_text=True,
                )

    # Column widths
    widths = {
        "A": 11,  # Nhóm
        "B": 13,  # Mã
        "C": 30,  # Tên
        "D": 40,  # Mô tả ngắn
        "E": 70,  # Mô tả chi tiết
        "F": 15,  # Role
        "G": 8,   # Phase
        "H": 11,  # Task
        "I": 16,  # Status
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    # Freeze first row
    ws.freeze_panes = "A2"

    # AutoFilter
    ws.auto_filter.ref = f"A1:I{len(FUNCTIONS) + 1}"

    return ws


# ──────────────────────────────────────────────────────────────────────
# Build Summary sheet
# ──────────────────────────────────────────────────────────────────────
def build_summary_sheet(wb):
    ws = wb.create_sheet("Tổng kết", 0)  # First sheet

    # Title
    ws.cell(row=1, column=1, value="Cura Clinic CMS — Function List Summary")
    ws.cell(row=1, column=1).font = Font(name="Calibri", size=18, bold=True, color=BRAND_BLUE)
    ws.merge_cells("A1:E1")

    ws.cell(row=2, column=1, value="Tổng kết phân loại theo nhóm + phase + status")
    ws.cell(row=2, column=1).font = Font(italic=True, color="666666")
    ws.merge_cells("A2:E2")

    # ─── Stats by status ────────────────────────────────────
    ws.cell(row=4, column=1, value="Theo trạng thái").font = Font(bold=True, size=14)
    status_counter = Counter(fn["status"] for fn in FUNCTIONS)
    headers = ["Status", "Count", "%"]
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=5, column=col_idx, value=h)
        c.font = Font(bold=True, color=HEADER_TEXT)
        c.fill = PatternFill("solid", fgColor=BRAND_BLUE)
        c.alignment = Alignment(horizontal="center")
        c.border = THIN_BORDER

    total = len(FUNCTIONS)
    row = 6
    for status in [DONE, WIP, TODO, IDEA]:
        count = status_counter.get(status, 0)
        pct = f"{count / total * 100:.1f}%" if total else "0.0%"
        ws.cell(row=row, column=1, value=status).fill = PatternFill("solid", fgColor=STATUS_COLORS[status])
        ws.cell(row=row, column=2, value=count).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=3, value=pct).alignment = Alignment(horizontal="center")
        for col_idx in range(1, 4):
            ws.cell(row=row, column=col_idx).border = THIN_BORDER
        row += 1
    ws.cell(row=row, column=1, value="TỔNG").font = Font(bold=True)
    ws.cell(row=row, column=2, value=total).font = Font(bold=True)
    ws.cell(row=row, column=2).alignment = Alignment(horizontal="center")
    ws.cell(row=row, column=3, value="100%").font = Font(bold=True)
    ws.cell(row=row, column=3).alignment = Alignment(horizontal="center")
    for col_idx in range(1, 4):
        ws.cell(row=row, column=col_idx).border = THIN_BORDER
        ws.cell(row=row, column=col_idx).fill = PatternFill("solid", fgColor="F0F0F0")

    # ─── Stats by phase ─────────────────────────────────────
    ws.cell(row=4, column=5, value="Theo phase").font = Font(bold=True, size=14)
    phase_counter = Counter(fn["phase"] for fn in FUNCTIONS)
    for col_idx, h in enumerate(headers, start=5):
        c = ws.cell(row=5, column=col_idx, value=h)
        c.font = Font(bold=True, color=HEADER_TEXT)
        c.fill = PatternFill("solid", fgColor=BRAND_BLUE)
        c.alignment = Alignment(horizontal="center")
        c.border = THIN_BORDER

    row = 6
    for phase in ["v1", "v2", "v3"]:
        count = phase_counter.get(phase, 0)
        pct = f"{count / total * 100:.1f}%" if total else "0.0%"
        ws.cell(row=row, column=5, value=phase).fill = PatternFill("solid", fgColor=PHASE_COLORS[phase])
        ws.cell(row=row, column=5).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=6, value=count).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=7, value=pct).alignment = Alignment(horizontal="center")
        for col_idx in range(5, 8):
            ws.cell(row=row, column=col_idx).border = THIN_BORDER
        row += 1

    # ─── Stats by group ─────────────────────────────────────
    ws.cell(row=12, column=1, value="Theo nhóm chức năng").font = Font(bold=True, size=14)

    group_headers = ["Nhóm", "Tổng", "DONE", "TODO", "IDEA", "v1", "v2", "v3", "% DONE"]
    for col_idx, h in enumerate(group_headers, start=1):
        c = ws.cell(row=13, column=col_idx, value=h)
        c.font = Font(bold=True, color=HEADER_TEXT)
        c.fill = PatternFill("solid", fgColor=BRAND_BLUE)
        c.alignment = Alignment(horizontal="center")
        c.border = THIN_BORDER

    # Build per-group stats
    groups_order = []
    seen = set()
    for fn in FUNCTIONS:
        if fn["group"] not in seen:
            seen.add(fn["group"])
            groups_order.append(fn["group"])

    row = 14
    for group in groups_order:
        items = [f for f in FUNCTIONS if f["group"] == group]
        g_total = len(items)
        g_done = sum(1 for f in items if f["status"] == DONE)
        g_todo = sum(1 for f in items if f["status"] == TODO)
        g_idea = sum(1 for f in items if f["status"] == IDEA)
        g_v1 = sum(1 for f in items if f["phase"] == "v1")
        g_v2 = sum(1 for f in items if f["phase"] == "v2")
        g_v3 = sum(1 for f in items if f["phase"] == "v3")
        g_pct = f"{g_done / g_total * 100:.0f}%" if g_total else "0%"

        cells = [group, g_total, g_done, g_todo, g_idea, g_v1, g_v2, g_v3, g_pct]
        for col_idx, val in enumerate(cells, start=1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.border = THIN_BORDER
            if col_idx == 1:
                cell.fill = PatternFill("solid", fgColor=GROUP_COLORS.get(group, "FFFFFF"))
                cell.font = Font(bold=True)
            else:
                cell.alignment = Alignment(horizontal="center")
        row += 1

    # Total row
    total_v1 = sum(1 for f in FUNCTIONS if f["phase"] == "v1")
    total_v2 = sum(1 for f in FUNCTIONS if f["phase"] == "v2")
    total_v3 = sum(1 for f in FUNCTIONS if f["phase"] == "v3")
    total_done = status_counter.get(DONE, 0)
    total_todo = status_counter.get(TODO, 0)
    total_idea = status_counter.get(IDEA, 0)
    total_pct = f"{total_done / total * 100:.0f}%"

    total_cells = ["TỔNG", total, total_done, total_todo, total_idea, total_v1, total_v2, total_v3, total_pct]
    for col_idx, val in enumerate(total_cells, start=1):
        cell = ws.cell(row=row, column=col_idx, value=val)
        cell.border = THIN_BORDER
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="F0F0F0")
        if col_idx > 1:
            cell.alignment = Alignment(horizontal="center")

    # Column widths
    widths = {"A": 11, "B": 8, "C": 8, "D": 8, "E": 8, "F": 8, "G": 8, "H": 8, "I": 10}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    # Legend section
    legend_row = row + 3
    ws.cell(row=legend_row, column=1, value="Chú giải Status").font = Font(bold=True, size=12)
    legend_row += 1
    legend_data = [
        (DONE, "Đã hoàn thành — code đã merge vào main, test pass"),
        (WIP, "Đang làm — task đang trong implementation/review/testing"),
        (TODO, "Chưa làm — đã spec nhưng chưa start"),
        (IDEA, "Ý tưởng — Phase 2/3, chưa spec chi tiết"),
    ]
    for status, desc in legend_data:
        ws.cell(row=legend_row, column=1, value=status).fill = PatternFill("solid", fgColor=STATUS_COLORS[status])
        ws.cell(row=legend_row, column=2, value=desc)
        ws.merge_cells(start_row=legend_row, start_column=2, end_row=legend_row, end_column=8)
        legend_row += 1

    # Footer
    legend_row += 2
    ws.cell(row=legend_row, column=1,
            value="Cập nhật: 2026-04-30 | Source: function_list_data.py").font = Font(italic=True, color="888888")
    ws.merge_cells(start_row=legend_row, start_column=1, end_row=legend_row, end_column=8)


# ──────────────────────────────────────────────────────────────────────
# Main
# ──────────────────────────────────────────────────────────────────────
def main():
    wb = openpyxl.Workbook()
    build_functions_sheet(wb)
    build_summary_sheet(wb)

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(f"OK Generated: {OUTPUT_PATH}")
    print(f"   Total functions: {len(FUNCTIONS)}")
    print(f"   Sheets: {wb.sheetnames}")


if __name__ == "__main__":
    main()
